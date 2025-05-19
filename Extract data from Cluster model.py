#!Data extraction step from Cluster results must be added here first!

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

file_path="U:/IESA-Opt-ModelLinking/data/20250430_detailed.xlsx"


def update_multiple_techs_and_years(file_path: str,sheet_name: str,tech_id_col: str,tech_id_row_start: int,merged_row: int,header_row: int,merged_name: str,update_data: dict, tech_ids_to_clear: list):
    """
        Clears values for all Tech_IDs in `tech_ids_to_clear`, then updates values for those in `update_data`.

        :param file_path: Path to Excel file
        :param sheet_name: Name of the worksheet
        :param tech_id_col: Column letter where Tech_IDs are located (e.g. 'A')
        :param tech_id_row_start: First row where Tech_IDs begin (e.g. 7)
        :param merged_row: Row with merged block title (e.g. 2)
        :param header_row: Row with year labels (e.g. 5)
        :param merged_name: Label of the merged header (e.g. 'Minimum use in a year')
        :param update_data: Dictionary of the form { Tech_ID: {year: value, ...}, ... }
        :param tech_ids_to_clear: List of All Tech_IDs whose values should be cleared first, to prevent mixing up results from different iterations.
        """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Step 1: Find merged column range
    merged_range = None
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= merged_row <= rng.max_row:
            if ws.cell(row=rng.min_row, column=rng.min_col).value == merged_name:
                merged_range = rng
                break
    if not merged_range:
        return f"❌ Merged header '{merged_name}' not found."

    # Step 2: Build mapping of year → column index
    year_to_column = {}
    for col in range(merged_range.min_col, merged_range.max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            try:
                year = int(str(val).strip())
                year_to_column[year] = col
            except ValueError:
                continue

    # Step 3: Build mapping of Tech_ID → row
    tech_col_index = column_index_from_string(tech_id_col)
    tech_id_to_row = {}
    row = tech_id_row_start
    while True:
        value = ws.cell(row=row, column=tech_col_index).value
        if value is None:
            break
        tech_id = str(value).strip()
        tech_id_to_row[tech_id] = row
        row += 1

    # Step 4: CLEAR values for all Tech_IDs in the clearing list
    for tech_id in tech_ids_to_clear:
        if tech_id not in tech_id_to_row:
            continue
        row_idx = tech_id_to_row[tech_id]
        for col in year_to_column.values():  # Only the year columns under the merged block
            cell = ws.cell(row=row_idx, column=col)
            cell.value = None  # Clear the cell

    # Step 5: Write values for each Tech_ID and year
    not_found = []
    for tech_id, year_vals in update_data.items():
        if tech_id not in tech_id_to_row:
            not_found.append(f"Tech_ID '{tech_id}' not found")
            continue
        for year, val in year_vals.items():
            col = year_to_column.get(year)
            if col is None:
                not_found.append(f"Year '{year}' not found under '{merged_name}'")
                continue
            ws.cell(row=tech_id_to_row[tech_id], column=col, value=val)

    # Save changes
    wb.save(file_path)

    if not_found:
        return "⚠️ Some entries were not found:\n" + "\n".join(not_found)
    return "✅ All values written successfully."

# Example usage
if __name__ == "__main__":
    updates = {
        "Res01_01": {2030: 10, 2050: 12.5},
        "Res01_02": {2030: 5.0, 2050: 7.5},
        "Res02_01": {2030: 5.0, 2050: 7.5}
    }
    tech_ids_to_clear = ["Res01_01","Res01_02", "Res02_01"]

    result = update_multiple_techs_and_years(
        file_path=file_path,
        sheet_name="Technologies",
        tech_id_col="A",
        tech_id_row_start=7,
        merged_row=2,
        header_row=5,
        merged_name="Minimum use in a year",
        update_data=updates,
        tech_ids_to_clear=tech_ids_to_clear
    )

    print(result)

